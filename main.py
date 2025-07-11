#!/usr/bin/env python3
"""
ClickHouse Daily Reporter
매일 오전 9:30에 실행되는 ClickHouse 쿼리 자동화 도구
"""

import os
import sys
import yaml
import logging
import pandas as pd
import subprocess
import time
import signal
from datetime import datetime
from pathlib import Path
import clickhouse_connect
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment

class ClickHouseReporter:
    def __init__(self, config_path="config.yaml"):
        self.config_path = config_path
        self.config = self.load_config()
        self.setup_logging()
        self.client = None
        self.port_forward_process = None
        self.cleanup_registered = False
        
    def load_config(self):
        """설정 파일 로드"""
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f)
        except FileNotFoundError:
            self.create_sample_config()
            print(f"설정 파일이 생성되었습니다: {self.config_path}")
            print("config.yaml 파일을 수정한 후 다시 실행해주세요.")
            sys.exit(1)
    
    def create_sample_config(self):
        """샘플 설정 파일 생성"""
        sample_config = {
            'clickhouse': {
                # 직접 연결 방식 (포트 포워딩 사용)
                'connection_type': 'direct',  # 'direct' 또는 'kubectl'
                'host': 'localhost',
                'port': 8123,
                'username': 'jhkim',
                'password': 'xxxx',
                'database': 'default',
                
                # kubectl 연결 방식 (k8s 환경)
                'kubectl': {
                    'enabled': True,
                    'pod_name': 'chi-signoz-clickhouse-cluster-0-0-0',
                    'namespace': 'clickhouse',
                    'internal_host': 'chi-signoz-clickhouse-cluster-1-0',
                    'internal_port': 8123,  # ClickHouse HTTP port
                    'port_forward_local_port': 8123,
                    'context': None  # kubectl context (None for default)
                }
            },
            'output': {
                'directory': './output',
                'filename_prefix': 'daily_report'
            },
            'queries': {
                'system_metrics': {
                    'name': '시스템 메트릭',
                    'query': '''
                    SELECT 
                        toDate(event_time) as date,
                        metric,
                        value
                    FROM system.metrics 
                    WHERE event_time >= today() - 1
                    ORDER BY event_time DESC
                    LIMIT 100
                    '''
                },
                'query_log': {
                    'name': '쿼리 로그',
                    'query': '''
                    SELECT 
                        toDate(event_time) as date,
                        query_duration_ms,
                        memory_usage,
                        query
                    FROM system.query_log 
                    WHERE event_time >= today() - 1
                    ORDER BY query_duration_ms DESC
                    LIMIT 50
                    '''
                }
            }
        }
        
        with open(self.config_path, 'w', encoding='utf-8') as f:
            yaml.dump(sample_config, f, default_flow_style=False, allow_unicode=True)
    
    def setup_logging(self):
        """로깅 설정"""
        log_dir = Path('logs')
        log_dir.mkdir(exist_ok=True)
        
        log_file = log_dir / f"reporter_{datetime.now().strftime('%Y%m%d')}.log"
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def connect_clickhouse(self):
        """ClickHouse 연결"""
        try:
            ch_config = self.config['clickhouse']
            connection_type = ch_config.get('connection_type', 'direct')
            
            if connection_type == 'kubectl':
                return self.connect_via_kubectl()
            else:
                return self.connect_direct()
                
        except Exception as e:
            self.logger.error(f"ClickHouse 연결 실패: {str(e)}")
            return False
    
    def connect_direct(self):
        """직접 연결 방식"""
        try:
            ch_config = self.config['clickhouse']
            self.client = clickhouse_connect.get_client(
                host=ch_config['host'],
                port=ch_config['port'],
                username=ch_config['username'],
                password=ch_config['password'],
                database=ch_config['database']
            )
            self.logger.info("ClickHouse 직접 연결 성공")
            return True
        except Exception as e:
            self.logger.error(f"ClickHouse 직접 연결 실패: {str(e)}")
            return False
    
    def connect_via_kubectl(self):
        """kubectl을 통한 연결 (포트 포워딩 사용)"""
        try:
            ch_config = self.config['clickhouse']
            kubectl_config = ch_config['kubectl']
            
            if not kubectl_config.get('enabled', False):
                self.logger.warning("kubectl 연결이 비활성화되어 있습니다. 직접 연결을 시도합니다.")
                return self.connect_direct()
            
            # kubectl 명령 가능 여부 확인
            if not self.check_kubectl_available():
                self.logger.error("kubectl 명령을 찾을 수 없습니다.")
                return False
            
            # 포트 포워딩 설정
            if not self.setup_port_forwarding():
                self.logger.error("포트 포워딩 설정 실패")
                return False
            
            # 정리 함수 등록
            if not self.cleanup_registered:
                import atexit
                atexit.register(self.cleanup_port_forwarding)
                signal.signal(signal.SIGTERM, self.signal_handler)
                signal.signal(signal.SIGINT, self.signal_handler)
                self.cleanup_registered = True
            
            # ClickHouse 클라이언트 연결
            local_port = kubectl_config.get('port_forward_local_port', 9000)
            self.client = clickhouse_connect.get_client(
                host='localhost',
                port=local_port,
                username=ch_config['username'],
                password=ch_config['password'],
                database=ch_config['database']
            )
            
            self.logger.info("ClickHouse kubectl 연결 성공")
            return True
            
        except Exception as e:
            self.logger.error(f"kubectl 연결 실패: {str(e)}")
            self.cleanup_port_forwarding()
            return False
    
    def check_kubectl_available(self):
        """kubectl 명령 사용 가능 여부 확인"""
        try:
            result = subprocess.run(['kubectl', 'version', '--client'], 
                                  capture_output=True, text=True, timeout=10)
            return result.returncode == 0
        except (subprocess.TimeoutExpired, FileNotFoundError):
            return False
    
    def setup_port_forwarding(self):
        """포트 포워딩 설정"""
        try:
            ch_config = self.config['clickhouse']
            kubectl_config = ch_config['kubectl']
            
            pod_name = kubectl_config['pod_name']
            namespace = kubectl_config['namespace']
            internal_port = kubectl_config.get('internal_port', 8123)
            local_port = kubectl_config.get('port_forward_local_port', 8123)
            context = kubectl_config.get('context')
            
            # kubectl context 설정
            if context:
                context_cmd = ['kubectl', 'config', 'use-context', context]
                self.logger.info(f"kubectl context 설정 중: {' '.join(context_cmd)}")
                
                result = subprocess.run(context_cmd, capture_output=True, text=True, timeout=30)
                if result.returncode != 0:
                    self.logger.error(f"kubectl context 설정 실패: {result.stderr}")
                    return False
                else:
                    self.logger.info(f"kubectl context 설정 성공: {context}")
            
            # kubectl 명령 구성
            cmd = ['kubectl', 'port-forward']
            if namespace:
                cmd.extend(['-n', namespace])
            
            cmd.extend([pod_name, f"{local_port}:{internal_port}"])
            
            self.logger.info(f"포트 포워딩 설정 중: {' '.join(cmd)}")
            
            # 포트 포워딩 프로세스 시작
            self.port_forward_process = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True
            )
            
            # 포트 포워딩이 준비될 때까지 대기
            for i in range(10):  # 최대 10초 대기
                if self.port_forward_process.poll() is not None:
                    # 프로세스가 종료됨
                    stdout, stderr = self.port_forward_process.communicate()
                    self.logger.error(f"포트 포워딩 실패: {stderr}")
                    return False
                
                # 포트 연결 테스트
                try:
                    import socket
                    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    sock.settimeout(1)
                    result = sock.connect_ex(('localhost', local_port))
                    sock.close()
                    if result == 0:
                        self.logger.info(f"포트 포워딩 준비 완료: localhost:{local_port}")
                        return True
                except:
                    pass
                
                time.sleep(1)
            
            self.logger.error("포트 포워딩 준비 시간 초과")
            return False
            
        except Exception as e:
            self.logger.error(f"포트 포워딩 설정 오류: {str(e)}")
            return False
    
    def cleanup_port_forwarding(self):
        """포트 포워딩 정리"""
        if self.port_forward_process and self.port_forward_process.poll() is None:
            self.logger.info("포트 포워딩 프로세스 종료 중...")
            self.port_forward_process.terminate()
            try:
                self.port_forward_process.wait(timeout=5)
            except subprocess.TimeoutExpired:
                self.port_forward_process.kill()
                self.port_forward_process.wait()
            self.logger.info("포트 포워딩 프로세스 종료 완료")
    
    def signal_handler(self, signum, frame):
        """시그널 핸들러"""
        self.logger.info(f"시그널 {signum} 수신, 정리 중...")
        self.cleanup_port_forwarding()
        sys.exit(0)
    
    def execute_query(self, query_name, query_info):
        """쿼리 실행"""
        try:
            self.logger.info(f"쿼리 실행 시작: {query_name}")
            
            result = self.client.query(query_info['query'])
            df = pd.DataFrame(result.result_rows, columns=result.column_names)
            
            self.logger.info(f"쿼리 완료: {query_name} - {len(df)} 행")
            return df
            
        except Exception as e:
            self.logger.error(f"쿼리 실행 실패 ({query_name}): {str(e)}")
            return None
    
    def create_excel_file(self, data_dict):
        """Excel 파일 생성"""
        try:
            # 출력 디렉토리 생성
            output_dir = Path(self.config['output']['directory'])
            output_dir.mkdir(exist_ok=True)
            
            # 파일명 생성
            date_str = datetime.now().strftime('%Y%m%d')
            filename = f"{self.config['output']['filename_prefix']}_{date_str}.xlsx"
            filepath = output_dir / filename
            
            # Excel 파일 생성
            wb = Workbook()
            wb.remove(wb.active)  # 기본 시트 제거
            
            for sheet_name, df in data_dict.items():
                if df is not None and not df.empty:
                    ws = wb.create_sheet(title=sheet_name[:31])  # Excel 시트명 길이 제한
                    
                    # 헤더 스타일 적용
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    # 데이터 추가
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=value)
                            
                            # 헤더 스타일 적용
                            if r_idx == 1:
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = Alignment(horizontal='center')
                    
                    # 열 너비 자동 조정
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
            
            # 파일 저장
            wb.save(filepath)
            self.logger.info(f"Excel 파일 생성 완료: {filepath}")
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"Excel 파일 생성 실패: {str(e)}")
            return None
    
    def run(self):
        """메인 실행 함수"""
        start_time = datetime.now()
        self.logger.info("=" * 50)
        self.logger.info(f"ClickHouse Daily Reporter 시작: {start_time}")
        
        try:
            # ClickHouse 연결
            if not self.connect_clickhouse():
                return False
            
            # 쿼리 실행
            results = {}
            queries = self.config.get('queries', {})
            
            for query_name, query_info in queries.items():
                df = self.execute_query(query_name, query_info)
                if df is not None:
                    sheet_name = query_info.get('name', query_name)
                    results[sheet_name] = df
            
            # Excel 파일 생성
            if results:
                filepath = self.create_excel_file(results)
                if filepath:
                    self.logger.info(f"리포트 생성 완료: {filepath}")
                    
                    # 실행 통계
                    end_time = datetime.now()
                    duration = end_time - start_time
                    self.logger.info(f"총 실행 시간: {duration}")
                    self.logger.info(f"총 {len(results)} 개 쿼리 실행 완료")
                    return True
            else:
                self.logger.warning("실행할 쿼리가 없거나 모든 쿼리가 실패했습니다.")
                return False
                
        except Exception as e:
            self.logger.error(f"실행 중 오류 발생: {str(e)}")
            return False
        finally:
            if self.client:
                self.client.close()
                self.logger.info("ClickHouse 연결 종료")
            
            # 포트 포워딩 정리
            self.cleanup_port_forwarding()
            
            end_time = datetime.now()
            self.logger.info(f"ClickHouse Daily Reporter 종료: {end_time}")
            self.logger.info("=" * 50)

def main():
    """메인 함수"""
    try:
        reporter = ClickHouseReporter()
        success = reporter.run()
        sys.exit(0 if success else 1)
    except KeyboardInterrupt:
        print("\n사용자에 의해 중단되었습니다.")
        sys.exit(1)
    except Exception as e:
        print(f"예상치 못한 오류가 발생했습니다: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()
    